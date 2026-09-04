import io
from decimal import Decimal
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    Member, PracticeType, Practice, Payment, Receipt,
    Expense, FinancialTransaction, AuditLog
)
from .serializers import (
    MemberSerializer, PracticeTypeSerializer, PracticeSerializer,
    PaymentSerializer, ReceiptSerializer, ExpenseSerializer,
    FinancialTransactionSerializer, AuditLogSerializer
)
from .services import FinancialService


class AuthViewSet(viewsets.ViewSet):
    permission_classes = [AllowAny]

    @action(detail=False, methods=['post'])
    def login(self, request):
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '').strip()

        if not username or not password:
            return Response({'success': False, 'error': 'يجب إدخال اسم المستخدم وكلمة المرور'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            login(request, user)
            return Response({
                'success': True,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name or 'المهندس المسؤول',
                    'email': user.email,
                    'is_staff': user.is_staff
                },
                'message': 'تم تسجيل الدخول بنجاح'
            })

        return Response({'success': False, 'error': 'اسم المستخدم أو كلمة المرور غير صحيحة'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def me(self, request):
        if request.user.is_authenticated:
            return Response({
                'authenticated': True,
                'user': {
                    'id': request.user.id,
                    'username': request.user.username,
                    'first_name': request.user.first_name or 'المهندس المسؤول',
                    'is_staff': request.user.is_staff
                }
            })
        return Response({'authenticated': False})

    @action(detail=False, methods=['post'])
    def logout(self, request):
        logout(request)
        return Response({'success': True, 'message': 'تم تسجيل الخروج بنجاح'})


class MemberViewSet(viewsets.ModelViewSet):
    serializer_class = MemberSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['full_name', 'mobile_number', 'national_id']

    def get_queryset(self):
        qs = Member.objects.filter(is_deleted=False).annotate(
            practices_count=Count('practices', filter=Q(practices__is_deleted=False), distinct=True)
        )
        street = self.request.query_params.get('street')
        is_active = self.request.query_params.get('is_active')
        search_query = self.request.query_params.get('search')

        if street:
            qs = qs.filter(street_number=street)
        if is_active is not None:
            qs = qs.filter(is_active=(is_active.lower() == 'true'))
        if search_query:
            qs = qs.filter(
                Q(full_name__icontains=search_query) |
                Q(mobile_number__icontains=search_query) |
                Q(national_id__icontains=search_query) |
                Q(id__iexact=search_query)
            )
        return qs.order_by('street_number', 'full_name')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if not instance.can_be_deleted():
            # Soft delete instead
            instance.is_deleted = True
            instance.is_active = False
            instance.save()
            AuditLog.objects.create(
                user=request.user if request.user.is_authenticated else None,
                action='MEMBER_ARCHIVED',
                entity_name='Member',
                entity_id=instance.id,
                old_values={'full_name': instance.full_name, 'is_deleted': False},
                new_values={'is_deleted': True}
            )
            return Response({'message': 'تم أرشفة العضو لحماية السجلات المالية السابقة'}, status=status.HTTP_200_OK)
        
        # Hard delete safe
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action='MEMBER_DELETED',
            entity_name='Member',
            entity_id=instance.id,
            old_values={'full_name': instance.full_name}
        )
        instance.delete()
        return Response({'message': 'تم حذف العضو بنجاح'}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'])
    def statement(self, request, pk=None):
        member = self.get_object()
        year = request.query_params.get('year', timezone.now().year)
        month = request.query_params.get('month', timezone.now().month)

        practices = member.practices.filter(is_deleted=False).order_by('-year', '-month')
        if year and month:
            curr_practices = practices.filter(year=year, month=month)
        else:
            curr_practices = practices

        total_req = sum(p.required_amount for p in curr_practices)
        total_paid = sum(p.total_paid for p in curr_practices)
        total_rem = sum(p.remaining_amount for p in curr_practices)
        total_over = sum(p.overpayment_amount for p in curr_practices)

        payments = member.payments.filter(is_voided=False).order_by('-payment_date', '-created_at')
        receipts = member.receipts.all().order_by('-practice__year', '-practice__month')

        return Response({
            'member': MemberSerializer(member).data,
            'period': {'year': year, 'month': month},
            'summary': {
                'total_required': str(total_req),
                'total_paid': str(total_paid),
                'remaining': str(total_rem),
                'overpayment': str(total_over),
                'status': 'FULLY_PAID' if (total_req > 0 and total_paid >= total_req) else 'UNPAID'
            },
            'practices': PracticeSerializer(practices, many=True).data,
            'payments': PaymentSerializer(payments, many=True).data,
            'receipts': ReceiptSerializer(receipts, many=True).data,
        })


class PracticeTypeViewSet(viewsets.ModelViewSet):
    queryset = PracticeType.objects.filter(is_active=True)
    serializer_class = PracticeTypeSerializer


class PracticeViewSet(viewsets.ModelViewSet):
    serializer_class = PracticeSerializer

    def get_queryset(self):
        qs = Practice.objects.filter(is_deleted=False).select_related('member', 'practice_type', 'receipt')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        member_id = self.request.query_params.get('member_id')
        street = self.request.query_params.get('street')
        search = self.request.query_params.get('search')

        if year:
            qs = qs.filter(year=year)
        if month:
            qs = qs.filter(month=month)
        if member_id:
            qs = qs.filter(member_id=member_id)
        if street:
            qs = qs.filter(member__street_number=street)
        if search:
            qs = qs.filter(
                Q(member__full_name__icontains=search) |
                Q(member__mobile_number__icontains=search) |
                Q(member__national_id__icontains=search)
            )
        return qs.order_by('member__street_number', 'member__full_name')

    def perform_create(self, serializer):
        practice = serializer.save()
        FinancialService.ensure_practice_receipt(practice)
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action='PRACTICE_CREATED',
            entity_name='Practice',
            entity_id=practice.id,
            new_values={
                'member_id': practice.member_id,
                'practice_type': practice.practice_type.name,
                'year': practice.year,
                'month': practice.month,
                'required_amount': str(practice.required_amount)
            }
        )

    @action(detail=False, methods=['post'])
    def bulk_create_month(self, request):
        """
        Quick bulk creation of practice for all active members for a specific month and year.
        """
        year = int(request.data.get('year', timezone.now().year))
        month = int(request.data.get('month', timezone.now().month))
        practice_type_id = request.data.get('practice_type_id')
        required_amount = Decimal(str(request.data.get('required_amount', '560.00')))
        
        practice_type = PracticeType.objects.get(id=practice_type_id)
        active_members = Member.objects.filter(is_active=True, is_deleted=False)

        created_count = 0
        with transaction.atomic():
            for m in active_members:
                practice, created = Practice.objects.get_or_create(
                    member=m,
                    practice_type=practice_type,
                    year=year,
                    month=month,
                    is_deleted=False,
                    defaults={'required_amount': required_amount}
                )
                if created:
                    FinancialService.ensure_practice_receipt(practice)
                    created_count += 1

        return Response({
            'success': True,
            'created_count': created_count,
            'message': f"تم إنشاء {created_count} ممارسة لشهر {month}/{year} بقيمة {required_amount} ج.م"
        })


class PaymentViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentSerializer

    def get_queryset(self):
        qs = Payment.objects.filter(is_voided=False).select_related('member', 'practice', 'practice__practice_type')
        member_id = self.request.query_params.get('member_id')
        practice_id = self.request.query_params.get('practice_id')
        street = self.request.query_params.get('street')
        
        if member_id:
            qs = qs.filter(member_id=member_id)
        if practice_id:
            qs = qs.filter(practice_id=practice_id)
        if street:
            qs = qs.filter(member__street_number=street)
            
        return qs.order_by('-payment_date', '-created_at')

    def create(self, request, *args, **kwargs):
        practice_id = request.data.get('practice')
        amount = request.data.get('amount')
        payment_date = request.data.get('payment_date')
        payment_method = request.data.get('payment_method', 'CASH')
        notes = request.data.get('notes', '')

        if not practice_id or not amount:
            return Response({'error': 'يجب تحديد الممارسة والمبلغ'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            payment = FinancialService.record_payment(
                practice_id=practice_id,
                amount=amount,
                payment_date=payment_date,
                payment_method=payment_method,
                notes=notes,
                user=request.user if request.user.is_authenticated else None,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            serializer = self.get_serializer(payment)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def void_payment(self, request, pk=None):
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response({'error': 'يجب توضيح سبب الإلغاء'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            payment = FinancialService.void_payment(
                payment_id=pk,
                reason=reason,
                user=request.user if request.user.is_authenticated else None,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return Response({'success': True, 'message': 'تم إلغاء الدفعة وعكس القيد المالي بنجاح'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ReceiptViewSet(viewsets.ModelViewSet):
    serializer_class = ReceiptSerializer

    def get_queryset(self):
        qs = Receipt.objects.select_related('member', 'practice', 'practice__practice_type')
        status_filter = self.request.query_params.get('status')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        street = self.request.query_params.get('street')
        search = self.request.query_params.get('search')

        if status_filter:
            qs = qs.filter(status=status_filter)
        if year:
            qs = qs.filter(practice__year=year)
        if month:
            qs = qs.filter(practice__month=month)
        if street:
            qs = qs.filter(member__street_number=street)
        if search:
            qs = qs.filter(
                Q(member__full_name__icontains=search) |
                Q(receipt_number__icontains=search) |
                Q(member__mobile_number__icontains=search)
            )

        return qs.order_by('member__street_number', 'member__full_name')

    @action(detail=True, methods=['post'])
    def mark_delivered(self, request, pk=None):
        receipt = self.get_object()
        receipt.status = 'DELIVERED'
        receipt.delivery_date = timezone.now().date()
        receipt.save()

        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action='RECEIPT_MARKED_DELIVERED',
            entity_name='Receipt',
            entity_id=receipt.id,
            new_values={'status': 'DELIVERED', 'delivery_date': str(receipt.delivery_date)}
        )
        return Response({'success': True, 'message': 'تم تسجيل تسليم الإيصال للعضو'})

    @action(detail=True, methods=['post'])
    def mark_received(self, request, pk=None):
        receipt = self.get_object()
        receipt.status = 'RECEIVED'
        receipt.received_date = timezone.now().date()
        receipt.save()

        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action='RECEIPT_MARKED_RECEIVED',
            entity_name='Receipt',
            entity_id=receipt.id,
            new_values={'status': 'RECEIVED', 'received_date': str(receipt.received_date)}
        )
        return Response({'success': True, 'message': 'تم تسجيل استلام الإيصال من شركة الكهرباء'})


class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer

    def get_queryset(self):
        qs = Expense.objects.filter(is_deleted=False)
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        if year:
            qs = qs.filter(expense_date__year=year)
        if month:
            qs = qs.filter(expense_date__month=month)
        return qs.order_by('-expense_date', '-created_at')

    def create(self, request, *args, **kwargs):
        title = request.data.get('title')
        amount = request.data.get('amount')
        expense_date = request.data.get('expense_date')
        payment_method = request.data.get('payment_method', 'CASH')
        description = request.data.get('description', '')
        document_image = request.FILES.get('document_image')

        try:
            expense = FinancialService.record_expense(
                title=title,
                amount=amount,
                expense_date=expense_date,
                payment_method=payment_method,
                description=description,
                document_image=document_image,
                user=request.user if request.user.is_authenticated else None,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return Response(self.get_serializer(expense).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class FinancialTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = FinancialTransactionSerializer

    def get_queryset(self):
        qs = FinancialTransaction.objects.select_related('member', 'practice', 'payment', 'expense')
        tx_type = self.request.query_params.get('type')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        member_id = self.request.query_params.get('member_id')

        if tx_type:
            qs = qs.filter(transaction_type=tx_type)
        if year:
            qs = qs.filter(transaction_date__year=year)
        if month:
            qs = qs.filter(transaction_date__month=month)
        if member_id:
            qs = qs.filter(member_id=member_id)

        return qs.order_by('-transaction_date', '-created_at')

    @action(detail=False, methods=['post'])
    def record_manual_overpayment(self, request):
        amount = request.data.get('amount')
        source_name = request.data.get('source_name', '').strip()
        payment_method = request.data.get('payment_method', 'CASH')
        description = request.data.get('description', '')
        transaction_date = request.data.get('transaction_date')

        if not amount or not source_name:
            return Response({'error': 'يجب تحديد المبلغ والجهة أو الشخص الدافع'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tx = FinancialService.record_manual_overpayment(
                amount=amount,
                source_name=source_name,
                transaction_date=transaction_date,
                payment_method=payment_method,
                description=description,
                user=request.user if request.user.is_authenticated else None,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return Response(self.get_serializer(tx).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related('user').order_by('-created_at')
    serializer_class = AuditLogSerializer


class ReportViewSet(viewsets.ViewSet):
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        total_members = Member.objects.filter(is_active=True, is_deleted=False).count()

        practices = Practice.objects.filter(year=year, month=month, is_deleted=False)
        total_required = practices.aggregate(total=Sum('required_amount'))['total'] or Decimal('0.00')

        # Real collections and overpayments for this month
        collections = FinancialTransaction.objects.filter(
            transaction_type='PRACTICE_COLLECTION',
            transaction_date__year=year,
            transaction_date__month=month
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        overpayments = FinancialTransaction.objects.filter(
            transaction_type='OVERPAYMENT',
            transaction_date__year=year,
            transaction_date__month=month
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        expenses = FinancialTransaction.objects.filter(
            transaction_type='EXPENSE',
            transaction_date__year=year,
            transaction_date__month=month
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        remaining = max(Decimal('0.00'), total_required - collections)
        net_balance = (collections + overpayments) - expenses

        # Status counts
        fully_paid_count = 0
        unpaid_count = 0

        for p in practices:
            st = p.payment_status
            if st == 'FULLY_PAID':
                fully_paid_count += 1
            else:
                unpaid_count += 1

        # Receipts status
        receipts = Receipt.objects.filter(practice__year=year, practice__month=month)
        receipts_received = receipts.filter(status='RECEIVED').count()
        receipts_delivered = receipts.filter(status='DELIVERED').count()
        receipts_not_received = receipts.filter(status='NOT_RECEIVED').count()

        return Response({
            'period': {'year': year, 'month': month},
            'members': {
                'total_members': total_members,
            },
            'collections': {
                'total_required': str(total_required),
                'total_paid': str(collections),
                'remaining': str(remaining),
            },
            'payment_status': {
                'fully_paid': fully_paid_count,
                'unpaid': unpaid_count,
            },
            'receipts': {
                'total': receipts.count(),
                'received': receipts_received,
                'delivered': receipts_delivered,
                'not_received': receipts_not_received,
            },
            'financials': {
                'overpayments': str(overpayments),
                'expenses': str(expenses),
                'net_balance': str(net_balance),
            }
        })

    @action(detail=False, methods=['get'])
    def streets(self, request):
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        streets_data = []
        for s in range(1, 18):
            members_count = Member.objects.filter(street_number=s, is_active=True, is_deleted=False).count()
            practices = Practice.objects.filter(member__street_number=s, year=year, month=month, is_deleted=False)
            req = sum(p.required_amount for p in practices)
            paid = sum(p.total_paid for p in practices)
            rem = sum(p.remaining_amount for p in practices)

            streets_data.append({
                'street_number': s,
                'street_name': f"شارع {s}",
                'members_count': members_count,
                'required_amount': str(req),
                'paid_amount': str(paid),
                'remaining_amount': str(rem),
                'practices_count': practices.count()
            })

        return Response({'year': year, 'month': month, 'streets': streets_data})

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        street = request.query_params.get('street')

        wb = openpyxl.Workbook()
        
        # Sheet 1: الممارسات والتحصيل
        ws1 = wb.active
        ws1.title = "كشف التحصيل والممارسات"
        ws1.views.sheetView[0].rightToLeft = True

        headers = ["م", "اسم العضو", "الشارع", "رقم الموبايل", "نوع الممارسة", "المطلوب", "المدفوع", "المتبقي", "المبلغ الزائد", "حالة السداد", "حالة الإيصال"]
        ws1.append(headers)

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        for col_num, cell in enumerate(ws1[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        practices = Practice.objects.filter(year=year, month=month, is_deleted=False).select_related('member', 'practice_type', 'receipt')
        if street:
            practices = practices.filter(member__street_number=street)

        row_idx = 1
        for p in practices:
            ws1.append([
                row_idx,
                p.member.full_name,
                f"شارع {p.member.street_number}",
                p.member.mobile_number,
                p.practice_type.name,
                float(p.required_amount),
                float(p.total_paid),
                float(p.remaining_amount),
                float(p.overpayment_amount),
                p.get_payment_status_display() if hasattr(p, 'get_payment_status_display') else p.payment_status,
                p.receipt.get_status_display() if hasattr(p, 'receipt') and p.receipt else 'بدون إيصال'
            ])
            row_idx += 1

        # Auto-adjust column widths
        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Green_Revolution_Report_{month}_{year}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
